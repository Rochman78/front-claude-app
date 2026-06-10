#!/usr/bin/env python3
"""Reporting SAV hebdomadaire — fetch Front API + génération Excel.

Convention (héritée du S22 du 02/06/2026) :
- Fenêtre affichée = 8 jours, du VENDREDI de la semaine précédente au VENDREDI de
  la semaine cible (inclus). Le Vendredi d'avant est "contextuel" (s'affiche avec
  traités = '—' et est exclu du TOTAL et de la cadence).
- Jours ouvrés comptés = Lun-Ven de la semaine cible, hors fériés FR.
- Heures équipe : Murella + Roniah, 7h/j sauf mardi 6h (réunion).
- Cadence prime individuelle = Murella et Roniah uniquement (Charles & Jérémy
  sont dans le débit équipe mais hors prime).

Usage :
    # Dernière semaine ISO complétée (défaut)
    python3 reporting_semaine.py

    # Semaine spécifique par numéro ISO
    python3 reporting_semaine.py --week 23

    # Dates explicites
    python3 reporting_semaine.py --start 2026-06-01 --end 2026-06-07

Sortie :
    ~/Downloads/reporting_S<NN>_2026.xlsx
    /tmp/reporting_S<NN>_raw.json  (réutilisable pour rebuild sans retaper l'API)
"""
import argparse, os, sys, json, time, re, random, datetime, subprocess
import urllib.request, urllib.parse, urllib.error
from zoneinfo import ZoneInfo

# ──────────────────────────────────────────────────────────────────────────
# CONFIG STATIQUE
# ──────────────────────────────────────────────────────────────────────────
PROJECT_ROOT = '/Users/charlesbamy/front-claude-app'
ENV = f'{PROJECT_ROOT}/.env'
BASE = 'https://api2.frontapp.com'
TZ = ZoneInfo('Europe/Paris')

TEAMMATES = {
    'tea_hmfvb': 'Murella',
    'tea_mxhsn': 'Roniah',
    'tea_mxqqf': 'Jeremy',
    'tea_gnazb': 'Zephyr',
}
TM_PRIME = ['tea_hmfvb', 'tea_mxhsn']

# Fériés FR (à étendre si besoin)
HOLIDAYS_FR = {
    '2026-01-01', '2026-04-06', '2026-05-01', '2026-05-08', '2026-05-14',
    '2026-05-25', '2026-07-14', '2026-11-11', '2026-12-25',
    '2027-01-01', '2027-03-29', '2027-05-01', '2027-05-06', '2027-05-08',
    '2027-05-17', '2027-07-14', '2027-08-15', '2027-11-01', '2027-11-11',
    '2027-12-25',
}

# Détection bruit reconstituée — peut différer 1-2 % du S22.
# À recalibrer si on observe des écarts importants.
NOISE_SENDER_RE = re.compile(
    r'(no-?reply|noreply|do-?not-?reply|mailer-?daemon|postmaster|'
    r'@(mailchimp|sendinblue|brevo|klaviyo|sendgrid|mailgun|amazonses|notify\.)|'
    r'bounces?@|newsletter@|notifications?@|alerts?@|invoices?@|facturation@|'
    r'octopia|cdiscount-marketplace)', re.I)
NOISE_SUBJECT_RE = re.compile(
    r'(out\s+of\s+office|absence|automatic\s+reply|automatique|vacation|'
    r'delivery\s+(status|failure)|undeliverable|mail\s+delivery|'
    r'newsletter|unsubscribe|d[ée]sabonn|\[OK\]|\[NEW\]|\[CRON\]|'
    r'facture\s+auto|invoice\s+#)', re.I)
SHOPIFY_LEGIT_RE = re.compile(
    r'(nouveau\s+message|klantbericht|nuevo\s+mensaje|neue\s+nachricht|'
    r'nuovo\s+messaggio|new\s+message)', re.I)

# ──────────────────────────────────────────────────────────────────────────
# UTILS DATE
# ──────────────────────────────────────────────────────────────────────────
def last_completed_iso_week(today=None):
    """Retourne (week_num, monday_date, friday_date) de la dernière semaine ISO
    complétée par rapport à `today`. Une semaine complétée = qui s'est terminée
    avant aujourd'hui."""
    today = today or datetime.date.today()
    # ISO week starts Monday. Today's ISO weekday: Mon=1..Sun=7
    iso_year, iso_week, iso_dow = today.isocalendar()
    # Monday of THIS week
    this_monday = today - datetime.timedelta(days=iso_dow - 1)
    # Last completed week's Monday = this Monday - 7 days
    last_monday = this_monday - datetime.timedelta(days=7)
    last_friday = last_monday + datetime.timedelta(days=4)
    _, last_week_num, _ = last_monday.isocalendar()
    return last_week_num, last_monday, last_friday

def week_window(monday, friday):
    """Fenêtre affichée = Ven précédent → Ven cible inclus (8 jours)."""
    pre_friday = monday - datetime.timedelta(days=3)   # Ven semaine d'avant
    end_exclusive = friday + datetime.timedelta(days=1)
    days = []
    d = pre_friday
    while d < end_exclusive:
        days.append(d.isoformat())
        d += datetime.timedelta(days=1)
    return days

def iso_week_to_monday(year, week):
    """Lundi d'une semaine ISO donnée."""
    jan4 = datetime.date(year, 1, 4)
    jan4_dow = jan4.isoweekday()
    week1_monday = jan4 - datetime.timedelta(days=jan4_dow - 1)
    return week1_monday + datetime.timedelta(weeks=week - 1)

# ──────────────────────────────────────────────────────────────────────────
# FRONT API
# ──────────────────────────────────────────────────────────────────────────
def load_token():
    return subprocess.check_output(
        f"grep '^FRONT_API_TOKEN=' {ENV} | cut -d= -f2-",
        shell=True
    ).decode().strip()

def get(url, token, max_retries=6):
    headers = {'Authorization': f'Bearer {token}', 'Accept': 'application/json'}
    if url.startswith('/'):
        url = BASE + url
    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=45) as r:
                return json.loads(r.read().decode())
        except urllib.error.HTTPError as e:
            if e.code == 429 or e.code >= 500:
                ra = e.headers.get('Retry-After')
                wait = float(ra) if ra else 2 ** attempt
                print(f'  ↻ HTTP {e.code}, wait {wait:.1f}s', file=sys.stderr, flush=True)
                time.sleep(wait); continue
            raise
        except (urllib.error.URLError, TimeoutError, OSError):
            time.sleep(2 ** attempt)
    raise RuntimeError(f'GET failed after retries: {url}')

def paginate(url, token):
    while url:
        data = get(url, token)
        yield from data.get('_results', [])
        url = (data.get('_pagination') or {}).get('next')

def is_noise(msg):
    if not msg.get('is_inbound'):
        return False
    sender = ''
    for rec in (msg.get('recipients') or []):
        if rec.get('role') == 'from':
            sender = (rec.get('handle') or '').lower(); break
    subj = (msg.get('subject') or '').lower()
    if 'shopify.com' in sender and SHOPIFY_LEGIT_RE.search(subj or ''):
        return False
    if sender and NOISE_SENDER_RE.search(sender): return True
    if subj and NOISE_SUBJECT_RE.search(subj): return True
    if 'shopify.com' in sender and not SHOPIFY_LEGIT_RE.search(subj or ''):
        return True
    return False

# ──────────────────────────────────────────────────────────────────────────
# FETCH
# ──────────────────────────────────────────────────────────────────────────
def fetch_data(days, token):
    start_dt = datetime.datetime.fromisoformat(days[0] + 'T00:00:00').replace(tzinfo=TZ)
    end_dt = datetime.datetime.fromisoformat(days[-1] + 'T00:00:00').replace(tzinfo=TZ) + datetime.timedelta(days=1)
    start_ts = start_dt.timestamp()
    end_ts = end_dt.timestamp()
    floor = start_ts - 86400

    print(f'[1/2] Fetch conversations dans la fenêtre {days[0]} → {days[-1]}…', flush=True)
    url = f'{BASE}/conversations?q[after]={int(floor)}&limit=100'
    convs = []
    seen = 0
    for c in paginate(url, token):
        seen += 1
        up = c.get('updated_at') or c.get('created_at') or 0
        if up < floor:
            break
        convs.append(c)
        if seen % 200 == 0:
            print(f'  · {seen} convs scannées, {len(convs)} retenues', file=sys.stderr, flush=True)
    print(f'  → {len(convs)} convs candidates\n', flush=True)

    rec_total = {d: 0 for d in days}
    rec_noise = {d: 0 for d in days}
    sent_tm = {tid: {d: 0 for d in days} for tid in TEAMMATES}
    comm_tm = {tid: {d: 0 for d in days} for tid in TEAMMATES}
    samples = {tid: [] for tid in TEAMMATES}

    print(f'[2/2] Fetch messages × {len(convs)} convs (single-thread, peut prendre 30-90 min selon le rate limit)…', flush=True)
    t0 = time.time()
    for idx, conv in enumerate(convs):
        cid = conv['id']
        try:
            msgs = list(paginate(f'{BASE}/conversations/{cid}/messages?limit=50', token))
        except Exception as e:
            print(f'  ✗ {cid}: {e}', file=sys.stderr); continue
        for m in msgs:
            ts = m.get('created_at') or 0
            if ts < start_ts or ts >= end_ts: continue
            day = datetime.datetime.fromtimestamp(ts, TZ).strftime('%Y-%m-%d')
            if day not in rec_total: continue
            if m.get('is_inbound'):
                rec_total[day] += 1
                if is_noise(m): rec_noise[day] += 1
            else:
                author = (m.get('author') or {}).get('id')
                if author in TEAMMATES:
                    sent_tm[author][day] += 1
                    samples[author].append({
                        'msg_id': m['id'], 'conv_id': cid, 'ts': ts,
                        'subject': m.get('subject') or conv.get('subject') or '',
                        'recipient': next((r.get('handle','') for r in (m.get('recipients') or [])
                                           if r.get('role')=='to'), ''),
                        'body': (m.get('text') or '')[:400],
                    })
        try:
            for c in paginate(f'{BASE}/conversations/{cid}/comments?limit=50', token):
                ts = c.get('posted_at') or 0
                if ts < start_ts or ts >= end_ts: continue
                day = datetime.datetime.fromtimestamp(ts, TZ).strftime('%Y-%m-%d')
                if day not in rec_total: continue
                author = (c.get('author') or {}).get('id')
                if author in TEAMMATES:
                    comm_tm[author][day] += 1
        except Exception:
            pass
        if (idx + 1) % 100 == 0:
            el = time.time() - t0
            eta = el / (idx+1) * (len(convs) - idx - 1)
            print(f'  · {idx+1}/{len(convs)} ({el:.0f}s, ETA {eta:.0f}s)', file=sys.stderr, flush=True)

    print(f'  ✓ done in {time.time()-t0:.0f}s\n')
    return {'days': days, 'rec_total': rec_total, 'rec_noise': rec_noise,
            'sent_tm': sent_tm, 'comm_tm': comm_tm, 'samples': samples}

# ──────────────────────────────────────────────────────────────────────────
# BUILD XLSX
# ──────────────────────────────────────────────────────────────────────────
def build_xlsx(raw, week_num, ctx_day, work_days, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    DAYS_FR = {0:'LUNDI',1:'MARDI',2:'MERCREDI',3:'JEUDI',4:'VENDREDI',5:'SAMEDI',6:'DIMANCHE'}
    days = raw['days']
    rec_total = raw['rec_total']; rec_noise = raw['rec_noise']
    sent_tm = raw['sent_tm']; comm_tm = raw['comm_tm']; samples = raw['samples']
    workset = set(work_days)

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = f'Reporting S{week_num}'

    bold = Font(bold=True, size=11)
    title = Font(bold=True, size=14, color='FFFFFF')
    title_fill = PatternFill('solid', fgColor='1F4E78')
    hdr_font = Font(bold=True, size=10, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='2E75B6')
    total_fill = PatternFill('solid', fgColor='FFE699')
    weekend_fill = PatternFill('solid', fgColor='EFEFEF')
    ctx_fill = PatternFill('solid', fgColor='F2F2F2')
    ferie_fill = PatternFill('solid', fgColor='F4CCCC')

    start_dmy = datetime.date.fromisoformat(days[0]).strftime('%d/%m')
    end_dmy = datetime.date.fromisoformat(days[-1]).strftime('%d/%m/%Y')
    ws.cell(row=1, column=1, value=f'REPORTING SAV — Semaine du {start_dmy} au {end_dmy}').font = title
    ws.cell(row=1, column=1).fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    ws.cell(row=3, column=1, value='PERFORMANCE INDIVIDUELLE').font = bold

    headers = ['JOUR','MAILS REÇUS','BRUITS','MAILS RÉELS','À TRAITER AUJ.','TOTAL TRAITÉS',
               'PAR MURELLA','PAR RONIAH','PAR JÉRÉMY','PAR ZEPHYR',
               'h TRAVAILLÉES ÉQUIPE','h MURELLA','h RONIAH',
               'CADENCE ÉQUIPE','CADENCE MURELLA','CADENCE RONIAH']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    row = 5
    cum_pre_workday = 0
    seen_workday = False
    prev_reels = None
    tot = {'rec':0,'noise':0,'reels':0,'traites':0,'M':0,'R':0,'J':0,'Z':0}

    for d in days:
        dt = datetime.date.fromisoformat(d)
        dow = dt.weekday()
        is_work = d in workset
        is_we = dow >= 5
        is_ferie = d in HOLIDAYS_FR
        is_ctx = d == ctx_day

        lbl = f'{DAYS_FR[dow]} {dt.strftime("%d/%m")}'
        if is_ctx: lbl += ' (S' + str(week_num - 1) + ')'
        if is_ferie: lbl += ' (férié)'
        ws.cell(row=row, column=1, value=lbl)

        rec = rec_total[d]; noise = rec_noise[d]; reels = rec - noise
        ws.cell(row=row, column=2, value=rec)
        ws.cell(row=row, column=3, value=noise)
        ws.cell(row=row, column=4, value=reels)

        if is_work:
            a_traiter = cum_pre_workday if not seen_workday else (prev_reels or 0)
            seen_workday = True
            ws.cell(row=row, column=5, value=a_traiter)
            tM = sent_tm['tea_hmfvb'][d] + comm_tm['tea_hmfvb'][d]
            tR = sent_tm['tea_mxhsn'][d] + comm_tm['tea_mxhsn'][d]
            tJ = sent_tm['tea_mxqqf'][d] + comm_tm['tea_mxqqf'][d]
            tZ = sent_tm['tea_gnazb'][d] + comm_tm['tea_gnazb'][d]
            tot_t = tM + tR + tJ + tZ
            ws.cell(row=row, column=6, value=tot_t)
            ws.cell(row=row, column=7, value=tM)
            ws.cell(row=row, column=8, value=tR)
            ws.cell(row=row, column=9, value=tJ)
            ws.cell(row=row, column=10, value=tZ)
            h_each = 6 if dow == 1 else 7
            h_eq = 2 * h_each
            ws.cell(row=row, column=11, value=h_eq)
            ws.cell(row=row, column=12, value=h_each)
            ws.cell(row=row, column=13, value=h_each)
            ws.cell(row=row, column=14, value=round(tot_t/h_eq, 1))
            ws.cell(row=row, column=15, value=round(tM/h_each, 1))
            ws.cell(row=row, column=16, value=round(tR/h_each, 1))
            tot['traites'] += tot_t; tot['M'] += tM; tot['R'] += tR
            tot['J'] += tJ; tot['Z'] += tZ
            prev_reels = reels
        else:
            ws.cell(row=row, column=5, value=0 if is_we else '—')
            for col in range(6, 17):
                ws.cell(row=row, column=col, value='—')
            cum_pre_workday += reels
            if is_we:
                for col in range(1, 17): ws.cell(row=row, column=col).fill = weekend_fill
            elif is_ctx:
                for col in range(1, 17): ws.cell(row=row, column=col).fill = ctx_fill
            elif is_ferie:
                for col in range(1, 17): ws.cell(row=row, column=col).fill = ferie_fill

        tot['rec'] += rec; tot['noise'] += noise; tot['reels'] += reels
        row += 1

    h_each_tot = sum(6 if datetime.date.fromisoformat(d).weekday() == 1 else 7 for d in work_days)
    h_eq_tot = 2 * h_each_tot

    ws.cell(row=row, column=1, value='TOTAL SEMAINE')
    ws.cell(row=row, column=2, value=tot['rec'])
    ws.cell(row=row, column=3, value=tot['noise'])
    ws.cell(row=row, column=4, value=tot['reels'])
    ws.cell(row=row, column=5, value='—')
    ws.cell(row=row, column=6, value=tot['traites'])
    ws.cell(row=row, column=7, value=tot['M'])
    ws.cell(row=row, column=8, value=tot['R'])
    ws.cell(row=row, column=9, value=tot['J'])
    ws.cell(row=row, column=10, value=tot['Z'])
    ws.cell(row=row, column=11, value=h_eq_tot)
    ws.cell(row=row, column=12, value=h_each_tot)
    ws.cell(row=row, column=13, value=h_each_tot)
    ws.cell(row=row, column=14, value=round(tot['traites']/h_eq_tot, 1) if h_eq_tot else 0)
    ws.cell(row=row, column=15, value=round(tot['M']/h_each_tot, 1) if h_each_tot else 0)
    ws.cell(row=row, column=16, value=round(tot['R']/h_each_tot, 1) if h_each_tot else 0)
    for col in range(1, 17):
        ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=col).font = bold
    row += 2

    ws.cell(row=row, column=1, value='SYNTHÈSE SEMAINE (BASE DE CALCUL CADENCE)').font = bold; row += 1
    work_start = datetime.date.fromisoformat(work_days[0]).strftime('%d/%m')
    work_end = datetime.date.fromisoformat(work_days[-1]).strftime('%d/%m')
    feries_in_week = [d for d in days if d in HOLIDAYS_FR]
    feries_label = f"dont {len(feries_in_week)} férié(s)" if feries_in_week else "aucun férié"
    for lbl, val in [
        (f'Période S{week_num}', f'{work_start} → {work_end}/{datetime.date.fromisoformat(work_days[-1]).year}'),
        ('Jours ouvrés', f'{len(work_days)} ({feries_label})'),
        ('Heures équipe (Mur+Ron)', f'{h_eq_tot}h ({h_each_tot}h × 2, mardi 6h pour réunion)'),
        ('Mails reçus brut (8 jours)', tot['rec']),
        ('Mails bruit filtrés', tot['noise']),
        ('Mails réels', tot['reels']),
        ('Mails traités équipe', tot['traites']),
        ('Cadence équipe', f'{round(tot["traites"]/h_eq_tot,1) if h_eq_tot else 0} mails/h'),
        ('Cadence Murella (prime)', f'{round(tot["M"]/h_each_tot,1) if h_each_tot else 0} mails/h sur {h_each_tot}h'),
        ('Cadence Roniah (prime)', f'{round(tot["R"]/h_each_tot,1) if h_each_tot else 0} mails/h sur {h_each_tot}h'),
        ('Note', 'Charles & Jérémy comptés dans le débit équipe mais HORS prime individuelle.'),
        ('Note', 'Filtre bruit reconstitué — peut différer 1-2 % du S22 originel.'),
        ('Note', f'Vendredi {ctx_day} affiché pour contexte uniquement (semaine précédente).'),
    ]:
        v = str(val); v = ' '+v if v.startswith('=') else v
        ws.cell(row=row, column=1, value=lbl)
        ws.cell(row=row, column=2, value=v)
        row += 1

    widths = [22,12,9,12,14,13,11,11,11,11,18,11,11,14,14,14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ─── Audit qualité ───
    ws2 = wb.create_sheet('Audit qualite')
    seed = int(datetime.date.today().strftime('%Y%m%d'))
    random.seed(seed)
    ws2.cell(row=1, column=1, value=f'AUDIT QUALITÉ — 5 mails par collaboratrice (notation /20)').font = bold
    ws2.cell(row=2, column=1, value=f'Échantillon stratifié reproductible (seed {seed}). Lien Front cliquable. Tirage sur jours ouvrés S{week_num}.')
    ws2.cell(row=3, column=1, value='Barème : ≥ 16/20 = prime entière (vert) | 10-15/20 = prime -50% (orange) | < 10/20 = prime 0 (rouge)')

    row = 5
    for tid, name in [('tea_hmfvb','Murella'), ('tea_mxhsn','Roniah')]:
        ws2.cell(row=row, column=1, value=f'■  {name}  ■').font = bold
        row += 1
        cols = ['#','Date envoi','Client','Sujet','Lien Front','Aperçu mail envoyé',
                'Compréhension /5','Exactitude /5','Ton /5','Résolution /5','TOTAL /20','Commentaire']
        for i, h in enumerate(cols, 1):
            c = ws2.cell(row=row, column=i, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1

        pool = [s for s in samples.get(tid, [])
                if datetime.datetime.fromtimestamp(s['ts'], TZ).strftime('%Y-%m-%d') in workset]
        by_day = {}
        for s in pool:
            day = datetime.datetime.fromtimestamp(s['ts'], TZ).strftime('%Y-%m-%d')
            by_day.setdefault(day, []).append(s)
        chosen = []
        for day in sorted(by_day):
            chosen.append(random.choice(by_day[day]))
        if len(chosen) > 5:
            chosen = random.sample(chosen, 5)
        elif len(chosen) < 5:
            rest = [s for s in pool if s not in chosen]
            chosen += random.sample(rest, min(5 - len(chosen), len(rest)))
        chosen.sort(key=lambda x: x['ts'])

        for i, s in enumerate(chosen, 1):
            dt = datetime.datetime.fromtimestamp(s['ts'], TZ)
            apercu = re.sub(r'\s+', ' ', s['body']).strip()[:300]
            if apercu.startswith('='): apercu = ' ' + apercu
            subj = (s['subject'] or '').strip()
            if subj.startswith('='): subj = ' ' + subj
            ws2.cell(row=row, column=1, value=i)
            ws2.cell(row=row, column=2, value=dt.strftime('%a %d/%m %H:%M'))
            ws2.cell(row=row, column=3, value=s['recipient'])
            ws2.cell(row=row, column=4, value=subj)
            c = ws2.cell(row=row, column=5, value='ouvrir →')
            c.hyperlink = f"https://app.frontapp.com/open/{s['conv_id']}"
            c.font = Font(color='0563C1', underline='single')
            ws2.cell(row=row, column=6, value=apercu)
            row += 1
        row += 2

    widths2 = [5, 16, 30, 60, 12, 80, 13, 12, 8, 13, 11, 30]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    return tot, h_eq_tot, h_each_tot

# ──────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--week', type=int, help="Numéro de semaine ISO (défaut = dernière complétée)")
    p.add_argument('--year', type=int, default=datetime.date.today().year)
    p.add_argument('--start', help="Date de début (YYYY-MM-DD) du lundi de la semaine cible")
    p.add_argument('--end', help="Date de fin (YYYY-MM-DD) du vendredi de la semaine cible")
    p.add_argument('--rebuild-only', action='store_true', help="Rebuild depuis raw JSON existant, sans retaper l'API")
    args = p.parse_args()

    if args.start and args.end:
        monday = datetime.date.fromisoformat(args.start)
        friday = datetime.date.fromisoformat(args.end)
        week_num = monday.isocalendar()[1]
    elif args.week:
        week_num = args.week
        monday = iso_week_to_monday(args.year, args.week)
        friday = monday + datetime.timedelta(days=4)
    else:
        week_num, monday, friday = last_completed_iso_week()

    days = week_window(monday, friday)
    ctx_day = days[0]   # Vendredi semaine précédente
    work_days = [d for d in days[1:]
                 if datetime.date.fromisoformat(d).weekday() < 5
                 and d not in HOLIDAYS_FR]

    print(f'═══ Reporting S{week_num} / {args.year} ═══')
    print(f'  Fenêtre affichée : {days[0]} → {days[-1]} (8 jours)')
    print(f'  Jours ouvrés     : {len(work_days)} ({work_days[0]} → {work_days[-1]})')
    print(f'  Sortie           : ~/Downloads/reporting_S{week_num}_{args.year}.xlsx')
    print()

    raw_path = f'/tmp/reporting_S{week_num}_{args.year}_raw.json'

    if args.rebuild_only:
        if not os.path.exists(raw_path):
            print(f'✗ raw JSON introuvable : {raw_path}', file=sys.stderr); sys.exit(1)
        print(f'  [rebuild-only] Lecture du raw : {raw_path}')
        with open(raw_path) as f:
            raw = json.load(f)
    else:
        token = load_token()
        raw = fetch_data(days, token)
        with open(raw_path, 'w') as f:
            json.dump(raw, f, indent=2, default=str)
        print(f'  raw → {raw_path}')

    out = os.path.expanduser(f'~/Downloads/reporting_S{week_num}_{args.year}.xlsx')
    tot, h_eq, h_each = build_xlsx(raw, week_num, ctx_day, work_days, out)

    print()
    print(f'═══ TOTAL SEMAINE S{week_num} ═══')
    print(f'  Mails réels (8j fenêtre)  : {tot["reels"]}')
    print(f'  Traités équipe            : {tot["traites"]}')
    print(f'    - Murella : {tot["M"]}')
    print(f'    - Roniah  : {tot["R"]}')
    print(f'    - Jérémy  : {tot["J"]}')
    print(f'    - Zephyr  : {tot["Z"]}')
    print(f'  Cadence équipe            : {round(tot["traites"]/h_eq, 1)} mails/h ({h_eq}h)')
    print(f'  Cadence Murella (prime)   : {round(tot["M"]/h_each, 1)} mails/h ({h_each}h)')
    print(f'  Cadence Roniah  (prime)   : {round(tot["R"]/h_each, 1)} mails/h ({h_each}h)')
    print()
    print(f'✓ {out}')

    # Auto-open
    try:
        subprocess.run(['open', out], check=False)
    except Exception:
        pass

if __name__ == '__main__':
    main()
